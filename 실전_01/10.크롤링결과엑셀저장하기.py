import requests
from bs4 import BeautifulSoup
import time
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 사용자입력
keyword = pyautogui.prompt("검색어를 입력하세요")
lastpage = int(pyautogui.prompt("몇 페이지까지 크롤링할까요?"))

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성하기
ws = wb.create_sheet(keyword)

# 열 너비 조절
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 120

# 행번호
row = 1

page_num = 1
for i in range(1,lastpage * 10, 10):
    print(f"{page_num}페이지크롤링 중입니다. ===========================>")
    response = requests.get(
        f"https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query={keyword}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    articles = soup.select("div.info_group")
    for article in articles:
        links = article.select("a.info")
        if len(links) >= 2:
            url = links[1].attrs["href"]
            print(url)
            response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            # 만약 연예뉴스 라면
            if "entertain" in response.url:
                title = soup.select_one(".end_tit")
                content = soup.select_one("#articeBody")
            elif "sports" in response.url:
                title = soup.select_one("h4.title")
                content = soup.select_one("#newsEndContents")
                # 본문 내용안에 불필요한 div 삭제
                divs = content.select("div")
                for div in divs:
                    div.decompose()
                paragraphs = content.select("p")
                for p in paragraphs:
                    p.decompose()
            else:
                title = soup.select_one("#title_area")
                content = soup.select_one("#dic_area")

            print("=============== 링크 ===============\n", url)
            print("=============== 제목 ===============\n", title.text.strip())
            print("=============== 본문 ===============\n", content.text.strip())

            ws[f'A{row}'] = url
            ws[f'B{row}'] = title.text.strip()
            ws[f'C{row}'] = content.text.strip()
            # 자동 줄바꿈
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)

            row = row + 1

            time.sleep(0.3)
    page_num = page_num + 1

wb.save(f"{keyword}_result.xlsx")